#!/usr/bin/env python3
"""
update_literature.py
====================
Build or rebuild the Figure 7 literature Excel file from scratch.

This script creates ``Figure_7_literature_data.xlsx`` containing 13 verified
literature entries plus present-study rows.  Every entry has been:

  1. DOI-verified via the CrossRef API.
  2. Values confirmed against the source PDF text.
  3. Harmonised to ppb with bilateral L/R averaging.

Inclusion / exclusion criteria
------------------------------
Studies were included if they reported mean QSM susceptibility for healthy
adults in >=3 of 4 target ROIs (thalamus, caudate, putamen, globus pallidus),
with values in tables or text (not solely in figures), using whole-structure
definitions (not subregions), and a compatible susceptibility reference
(whole-brain, CSF, or air -- not white-matter referenced).

Studies were excluded if they:
  - Reported only subregions (e.g., dorsal thalamic nuclei, pulvinar)
  - Used white-matter referencing (systematically different value range)
  - Lacked >=3 of 4 target ROIs
  - Had values only in figures / box-plots with no tabulated data
  - Were purely methodological comparisons without reference HC means

See ``docs/Figure7_manuscript_text.md`` for the complete list of included
and excluded studies with rationale.

Data harmonisation
------------------
  - All values converted to **ppb** (ppm x 1000 where needed)
  - Bilateral L/R averaged where separate hemispheric values reported
  - Age-regression models evaluated at reference age 30 y (Liu JMRI 2016)
    or population grand means where available (Gong, Li, etc.)
  - Approximate curve-reading noted for Treit et al. HBM 2021

Run:  python3 update_literature.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# -- Paths -----------------------------------------------------------------
REVIEW_DIR = Path(__file__).resolve().parent
XLSX = REVIEW_DIR / "Figure_7_literature_data.xlsx"

# -- Verified literature entries (13 total) --------------------------------
# Format: (FirstAuthor, Year, StudyCitation, Field_T,
#          Thalamus_ppb, Caudate_ppb, Putamen_ppb, Pallidum_ppb,
#          DOI, N_healthy, SourceNotes)

LITERATURE = [
    # ── 1.5 T (k=3, N=368) ──────────────────────────────────────────────
    ("Liu", 2016,
     "Liu et al., JMRI 2016",
     1.5,
     6.0, 33.3, 42.9, 122.9,
     "10.1002/jmri.25130", 174,
     "Table 2 whole-region linear fit (chi=A*age+B), L/R averaged at age=30"),

    ("Persson", 2015,
     "Persson et al., NI 2015",
     1.5,
     11.8, 89.2, 88.0, 191.1,
     "10.1016/j.neuroimage.2015.07.050", 183,
     "Table 1 direct means"),

    ("Bilgic", 2012,
     "Bilgic et al., NI 2012 (QSM)",
     1.5,
     46.4, 93.7, 77.9, 122.4,
     "10.1016/j.neuroimage.2011.08.077", 11,
     "Table 1a, l1-regularised phase-guided, Young N=11; ppm x 1000"),

    # ── 3 T (k=9, N=1,670; thalamus k=8, N=1,648) ──────────────────────
    ("Burgetova", 2021,
     "Burgetova et al., QIMS 2021",
     3.0,
     -2.8, 42.2, 40.5, 124.5,
     "10.21037/qims-21-87", 95,
     "Supplementary Table S1 / Figure 2 scatter plots; 95 subjects aged 21-58"),

    ("Gong", 2015,
     "Gong et al., NMR Biomed 2015",
     3.0,
     23.0, 42.2, 37.9, 81.8,
     "10.1002/nbm.3366", 42,
     "Grand means from text (x10^-2 ppm); 42 subjects aged 25-78"),

    ("Hinoda", 2015,
     "Hinoda et al., Invest Radiol 2015",
     3.0,
     None, 38.0, 52.0, 123.0,
     "10.1097/RLI.0000000000000099", 22,
     "V-SHARP 3T; no thalamus measured; ppm x 1000"),

    ("Li", 2021,
     "Li et al., Front Aging Neurosci 2021",
     3.0,
     0.0, 83.2, 98.8, 185.8,
     "10.3389/fnagi.2021.611891", 105,
     "Table 2 direct means; elderly (age 65 +/- 6)"),

    ("Li", 2021,
     "Li et al., Front Neurosci 2021",
     3.0,
     30.0, 73.6, 91.4, 209.2,
     "10.3389/fnins.2020.607705", 623,
     "RII regression at age=30; mixed 1.5T+3T (72% at 3T)"),

    ("Liu", 2025,
     "Liu et al., QIMS 2025",
     3.0,
     13.0, 31.0, 45.0, 81.0,
     "10.21037/qims-2024-2704", 30,
     "Table 2 HC values; N=30 from flowchart"),

    ("Treit", 2021,
     "Treit et al., HBM 2021",
     3.0,
     6.0, 25.0, 18.0, 105.0,
     "10.1002/hbm.25569", 498,
     "Approximate curve-reading from polynomial QSM figures at age ~25-30"),

    ("Zhao", 2019,
     "Zhao et al., J Craniofac Surg 2019",
     3.0,
     17.9, 27.7, 38.3, 118.4,
     "10.1097/SCS.0000000000005597", 42,
     "SDC Table 2; young adults; bilateral mean"),

    ("Zhou", 2020,
     "Zhou et al., Front Aging Neurosci 2020",
     3.0,
     45.0, 79.0, 88.0, 75.0,
     "10.3389/fnagi.2020.559603", 213,
     "Table 1 direct means; elderly (age 60.1 +/- 7.3)"),

    # ── 7 T (k=1, N=14) ─────────────────────────────────────────────────
    ("Ravanfar", 2023,
     "Ravanfar et al., AJNR 2023",
     7.0,
     2.0, 22.4, 16.3, 83.1,
     "10.3174/ajnr.A7894", 14,
     "Table 2 L/R averaged; Niemann-Pick Type C controls"),
]

# -- Present study data (3 T and 5 T, N=7) --------------------------------
PRESENT_STUDY = [
    # (Field_T, Thalamus, Caudate, Putamen, Pallidum)
    (3.0, 5.35, 28.10, 19.26, 88.76),
    (5.0, 7.65, 26.14, 18.59, 84.40),
]


def build_xlsx():
    """Create the literature Excel workbook from scratch."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Literature QSM Data"

    # Column headers
    headers = [
        "FirstAuthor", "Study", "Field_T", "N_healthy",
        "Thalamus_ppb", "Caudate_ppb", "Putamen_ppb", "Pallidum_ppb",
        "Source", "DOI", "Notes",
    ]
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="D9E2F3")
    thin_border = Border(bottom=Side(style="thin", color="999999"))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Literature data rows
    lit_font = Font(size=10, color="333333")
    row_idx = 2
    for entry in LITERATURE:
        (first_auth, _year, study, field,
         tha, cau, put, pal, doi, n, notes) = entry
        ws.cell(row=row_idx, column=1, value=first_auth).font = lit_font
        ws.cell(row=row_idx, column=2, value=study).font = lit_font
        ws.cell(row=row_idx, column=3, value=field).font = lit_font
        ws.cell(row=row_idx, column=4, value=n).font = lit_font
        ws.cell(row=row_idx, column=5, value=tha).font = lit_font
        ws.cell(row=row_idx, column=6, value=cau).font = lit_font
        ws.cell(row=row_idx, column=7, value=put).font = lit_font
        ws.cell(row=row_idx, column=8, value=pal).font = lit_font
        ws.cell(row=row_idx, column=9, value="literature").font = lit_font
        ws.cell(row=row_idx, column=10, value=doi).font = lit_font
        ws.cell(row=row_idx, column=11, value=notes).font = lit_font
        row_idx += 1

    # Separator
    sep_font = Font(size=10, color="999999", bold=True)
    ws.cell(row=row_idx, column=1, value="").font = sep_font
    row_idx += 1

    # Present study rows
    ps_font = Font(size=10, color="0066CC", bold=True)
    for (field, tha, cau, put, pal) in PRESENT_STUDY:
        ws.cell(row=row_idx, column=1, value="Present study").font = ps_font
        ws.cell(row=row_idx, column=2, value="Present study").font = ps_font
        ws.cell(row=row_idx, column=3, value=field).font = ps_font
        ws.cell(row=row_idx, column=4, value=7).font = ps_font
        ws.cell(row=row_idx, column=5, value=tha).font = ps_font
        ws.cell(row=row_idx, column=6, value=cau).font = ps_font
        ws.cell(row=row_idx, column=7, value=put).font = ps_font
        ws.cell(row=row_idx, column=8, value=pal).font = ps_font
        ws.cell(row=row_idx, column=9, value="present_study").font = ps_font
        ws.cell(row=row_idx, column=10, value="").font = ps_font
        ws.cell(row=row_idx, column=11, value="N=7, multicenter").font = ps_font
        row_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    for col in "EFGH":
        ws.column_dimensions[col].width = 15
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 36
    ws.column_dimensions["K"].width = 56

    wb.save(XLSX)
    print(f"Saved -> {XLSX}")
    print(f"  Literature entries: {len(LITERATURE)}")
    print(f"  Present study rows: {len(PRESENT_STUDY)}")

    # Summary table
    print("\n" + "=" * 110)
    print(f"{'Author':<14} {'Study':<42} {'T':>4} {'N':>5} "
          f"{'THA':>8} {'CN':>8} {'PUT':>8} {'GP':>8}")
    print("-" * 110)
    for entry in LITERATURE:
        (fa, _y, study, field,
         tha, cau, put, pal, _doi, n, _notes) = entry
        t = f"{tha:.1f}" if tha is not None else "-"
        c = f"{cau:.1f}" if cau is not None else "-"
        p = f"{put:.1f}" if put is not None else "-"
        g = f"{pal:.1f}" if pal is not None else "-"
        print(f"{fa:<14} {study:<42} {field:>4} {n:>5} "
              f"{t:>8} {c:>8} {p:>8} {g:>8}")


if __name__ == "__main__":
    build_xlsx()

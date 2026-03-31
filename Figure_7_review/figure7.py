#!/usr/bin/env python3
"""
regenerate_figure7.py
=====================
Regenerate Figure 7 from the verified literature Excel file.

This script reads literature QSM reference data and present-study values
from an Excel workbook, then produces a four-panel figure showing
susceptibility across magnetic field strengths for four deep gray-matter
ROIs (thalamus, caudate nucleus, putamen, globus pallidus).

Literature database: **13 verified entries** (all confirmed against source
PDFs, DOIs verified via CrossRef API).

  1.5 T -- 3 studies,  N = 368
    3 T -- 9 studies,  N = 1,648-1,670  (8 for thalamus; Hinoda lacks THA)
    7 T -- 1 study,    N = 14

Inclusion criteria:
  - Healthy adult controls; >=3 of 4 target ROIs reported
  - Whole-structure values (not subregions)
  - Compatible susceptibility reference (not WM-referenced)
  - Numerical values in tables/text (not solely in figures)

Data harmonisation:
  - All values in ppb; bilateral L/R averaged
  - Age-regression models evaluated at age 30 (or population centre)
  - See docs/Figure7_manuscript_text.md for full details

Removed entries:
  - 7 entries with invalid/unverifiable DOIs
  - Deistung et al. NI 2013 (wrong N, WM-referenced, subregion thalamus)

Run:  python3 regenerate_figure7.py
"""

from pathlib import Path
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import matplotlib.patches as mpatches
from scipy.interpolate import make_interp_spline
import openpyxl

# -- Paths ----------------------------------------------------------------
PROJECT = Path(__file__).resolve().parent.parent
FIG7_DIR = PROJECT / 'output' / 'Figure7'
REVIEW_DIR = Path(__file__).resolve().parent
XLSX = REVIEW_DIR / 'Figure_7_literature_data.xlsx'

# -- Colour palette --------------------------------------------------------
_cmap = matplotlib.colormaps.get_cmap('RdBu_r')
_norm = matplotlib.colors.Normalize(vmin=-50, vmax=50)
def _c(ppb):
    return matplotlib.colors.rgb2hex(_cmap(_norm(ppb))[:3])

C5T = _c(40)
lit_color  = '#2166AC'
lit_band   = '#92C5DE'
our_color  = C5T
our_alpha  = 0.70

# -- ROIs ------------------------------------------------------------------
ROIS = ['Thalamus', 'Caudate', 'Putamen', 'Pallidum']

# -- Study subjects (for k / N annotations) --------------------------------
# Final verified set -- 13 literature entries
_STUDY_SUBJECTS = {
    # 1.5 T (k=3, N=368)
    ('Liu et al., JMRI 2016',                1.5): 174,
    ('Persson et al., NI 2015',              1.5): 183,
    ('Bilgic et al., NI 2012 (QSM)',         1.5):  11,
    # 3 T (k=9, N=1670; thalamus k=8, N=1648)
    ('Burgetova et al., QIMS 2021',          3.0):  95,
    ('Gong et al., NMR Biomed 2015',         3.0):  42,
    ('Hinoda et al., Invest Radiol 2015',    3.0):  22,   # no thalamus
    ('Li et al., Front Aging Neurosci 2021', 3.0): 105,
    ('Li et al., Front Neurosci 2021',       3.0): 623,
    ('Liu et al., QIMS 2025',                3.0):  30,
    ('Treit et al., HBM 2021',              3.0): 498,
    ('Zhao et al., J Craniofac Surg 2019',   3.0):  42,
    ('Zhou et al., Front Aging Neurosci 2020', 3.0): 213,
    # 7 T (k=1, N=14)
    ('Ravanfar et al., AJNR 2023',           7.0):  14,
}

# Studies that are MISSING specific ROIs
_MISSING_ROIS = {
    'Hinoda et al., Invest Radiol 2015': {'Thalamus'},
}


def read_xlsx():
    """Read the literature Excel file."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active

    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    new_fmt = (hdr[0] == 'FirstAuthor')

    lit_rows = []
    our_data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        while len(vals) < 11:
            vals.append(None)

        if new_fmt:
            _fa, study, field, _n, tha, cau, put, pal, src = vals[:9]
        else:
            study, field, tha, cau, put, pal, src = vals[:7]

        if study is None or field is None:
            continue
        if str(study).startswith('--'):
            continue

        if src == 'present_study':
            field = float(field)
            for i, roi in enumerate(ROIS):
                col_val = [tha, cau, put, pal][i]
                if col_val is not None:
                    our_data.setdefault(roi, {})[field] = float(col_val)
            continue

        lit_rows.append((study, float(field), tha, cau, put, pal))

    wb.close()
    return lit_rows, our_data


def smooth_curve(field_vals, y_vals, n_pts=200):
    x = np.array(field_vals, dtype=float)
    y = np.array(y_vals, dtype=float)
    if len(x) < 3:
        x_s = np.linspace(x[0], x[-1], n_pts)
        y_s = np.interp(x_s, x, y)
        return x_s, y_s
    k = min(3, len(x) - 1)
    spl = make_interp_spline(x, y, k=k)
    x_s = np.linspace(x[0], x[-1], n_pts)
    y_s = spl(x_s)
    return x_s, y_s


def compute_per_roi_stats(lit_rows):
    """Compute per-ROI k (paper count) and N (total subjects)."""
    lit = {roi: {} for roi in ROIS}
    for study, field, tha, cau, put, pal in lit_rows:
        for roi, val in zip(ROIS, [tha, cau, put, pal]):
            if val is not None and val != '':
                lit[roi].setdefault(field, []).append(float(val))

    lit_summary = {}
    for roi in ROIS:
        lit_summary[roi] = {}
        for field in sorted(lit[roi]):
            vals = lit[roi][field]
            lit_summary[roi][field] = (np.mean(vals), min(vals), max(vals))

    roi_k = {roi: {} for roi in ROIS}
    roi_N = {roi: {} for roi in ROIS}

    for roi_idx, roi in enumerate(ROIS):
        for study, field, tha, cau, put, pal in lit_rows:
            val = [tha, cau, put, pal][roi_idx]
            if val is None or val == '':
                continue
            roi_k[roi].setdefault(field, set()).add(study)
            n_subj = _STUDY_SUBJECTS.get((study, field))
            if n_subj is not None:
                if study not in roi_N[roi].get(field, {}).get('_seen', set()):
                    roi_N[roi].setdefault(field, {'N': 0, '_seen': set()})
                    roi_N[roi][field]['N'] += n_subj
                    roi_N[roi][field]['_seen'].add(study)

    k_papers = {roi: {f: len(s) for f, s in roi_k[roi].items()} for roi in ROIS}
    N_subjects = {roi: {f: roi_N[roi].get(f, {}).get('N', 0)
                        for f in roi_k[roi]} for roi in ROIS}

    return lit, lit_summary, k_papers, N_subjects


def main():
    print(f"Reading: {XLSX}")
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found.")
        return

    lit_rows, our_data = read_xlsx()
    print(f"  Literature rows: {len(lit_rows)}")
    print(f"  Present study ROIs: {list(our_data.keys())}")

    lit, lit_summary, k_papers, N_subjects = compute_per_roi_stats(lit_rows)

    for roi in ROIS:
        print(f"\n  {roi}:")
        for f in sorted(k_papers[roi]):
            print(f"    {f}T: k={k_papers[roi][f]}, N={N_subjects[roi].get(f, '?')}")

    # -- Create figure -------------------------------------------------------
    fig, axes = plt.subplots(1, 4, figsize=(18, 5.0), sharey=False)
    fig.suptitle('QSM Susceptibility Across Magnetic Field Strengths',
                 fontsize=14, fontweight='bold', y=0.98)

    panel_labels = ['A', 'B', 'C', 'D']

    for ax_idx, (ax, roi) in enumerate(zip(axes, ROIS)):

        fields_sorted = sorted(lit_summary[roi])
        means = [lit_summary[roi][f][0] for f in fields_sorted]
        lo    = [lit_summary[roi][f][1] for f in fields_sorted]
        hi    = [lit_summary[roi][f][2] for f in fields_sorted]

        xs_m, ys_m = smooth_curve(fields_sorted, means)
        xs_l, ys_l = smooth_curve(fields_sorted, lo)
        xs_h, ys_h = smooth_curve(fields_sorted, hi)

        ax.fill_between(xs_m, ys_l, ys_h, color=lit_band,
                        alpha=0.35, zorder=1, linewidth=0)
        ax.plot(xs_m, ys_m, color=lit_color, linewidth=2.8,
                alpha=0.90, zorder=3)
        ax.scatter(fields_sorted, means, s=40, color=lit_color,
                   edgecolors='white', linewidth=0.8, zorder=4)

        # Per-ROI k, N annotations
        for f_pos, m_pos in zip(fields_sorted, means):
            k = k_papers[roi].get(f_pos, 0)
            N = N_subjects[roi].get(f_pos, 0)
            if k > 0 and N > 0:
                if roi == 'Pallidum' and f_pos == 3.0:
                    y_off, va = 12, 'bottom'
                else:
                    y_off, va = -14, 'top'
                ax.annotate(f'k={k}, N={N}',
                            xy=(f_pos, m_pos),
                            xytext=(0, y_off), textcoords='offset points',
                            fontsize=6.5, color=lit_color, ha='center',
                            va=va, fontstyle='italic')

        # Present study curve (3 T -> 5 T)
        our_fields = sorted(our_data.get(roi, {}))
        our_means = [our_data[roi][f] for f in our_fields]
        if len(our_fields) >= 2:
            ax.plot(our_fields, our_means, color=our_color,
                    linewidth=2.0, linestyle='-', alpha=our_alpha,
                    zorder=5)
        for f_T, m_val in zip(our_fields, our_means):
            ax.scatter(f_T, m_val, s=60, color=our_color,
                       edgecolors='white', linewidth=1.0,
                       alpha=our_alpha, zorder=6)

        # Present study N annotation
        for f_T, m_val in zip(our_fields, our_means):
            if roi in ('Caudate', 'Putamen'):
                y_off, va = -14, 'top'
            else:
                y_off, va = 10, 'bottom'
            ax.annotate('N=7',
                        xy=(f_T, m_val),
                        xytext=(0, y_off), textcoords='offset points',
                        fontsize=6.5, color=our_color, ha='center',
                        va=va, fontstyle='italic')

        # Formatting
        ax.set_xlim(0.8, 7.8)
        ax.set_xticks([1.5, 3, 5, 7])
        ax.set_xticklabels(['1.5', '3', '5', '7'], fontsize=10)
        ax.set_xlabel('Field Strength (T)', fontsize=10)
        if ax_idx == 0:
            ax.set_ylabel('Susceptibility (ppb)', fontsize=10)
        ax.set_title(f'{panel_labels[ax_idx]}   {roi}',
                     fontsize=12, fontweight='bold', loc='left')
        ax.tick_params(labelsize=9)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.axhline(y=0, color='#CCCCCC', linewidth=0.7, linestyle='--',
                   zorder=0)
        ax.axvspan(4.3, 5.7, alpha=0.04, color=our_color, zorder=0)

    # Legend
    h_lit = Line2D([], [], color=lit_color, linewidth=2.8, alpha=0.9,
                   marker='o', markersize=5, markeredgecolor='white',
                   markeredgewidth=0.8,
                   label='Literature consensus (1.5\u20137 T)')
    h_band = mpatches.Patch(color=lit_band, alpha=0.35,
                            label='Inter-study range')
    h_our = Line2D([], [], color=our_color, linewidth=2.0, alpha=our_alpha,
                   marker='o', markersize=5, markeredgecolor='white',
                   markeredgewidth=0.8,
                   label='Present study (3 T \u2192 5 T)')
    fig.legend(handles=[h_lit, h_band, h_our], loc='upper center',
               bbox_to_anchor=(0.5, 0.935), ncol=3,
               fontsize=9, frameon=True, framealpha=0.92,
               edgecolor='#CCCCCC', handletextpad=0.5,
               columnspacing=2.0, borderpad=0.5)

    plt.tight_layout(rect=[0, 0, 1, 0.87])

    FIG7_DIR.mkdir(parents=True, exist_ok=True)

    out_pdf = FIG7_DIR / 'Figure_7_final.pdf'
    fig.savefig(out_pdf, dpi=300, bbox_inches='tight')
    out_png = FIG7_DIR / 'Figure_7_final.png'
    fig.savefig(out_png, dpi=300, bbox_inches='tight')
    plt.close(fig)

    print(f"\nSaved -> {out_pdf}")
    print(f"Saved -> {out_png}")

    # Also copy to top-level output
    import shutil
    top_pdf = PROJECT / 'output' / 'Figure_7_final.pdf'
    shutil.copy2(out_pdf, top_pdf)
    print(f"Copied -> {top_pdf}")

    # CSV summary
    import csv
    csv_path = FIG7_DIR / 'Figure_7_data_final.csv'
    with open(csv_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=[
            'ROI', 'source', 'field_T', 'mean_ppb', 'k_papers', 'N_subjects'])
        writer.writeheader()
        for roi in ROIS:
            for field in sorted(lit_summary[roi]):
                m, lo, hi = lit_summary[roi][field]
                writer.writerow({
                    'ROI': roi, 'source': 'literature',
                    'field_T': field, 'mean_ppb': round(m, 2),
                    'k_papers': k_papers[roi].get(field, 0),
                    'N_subjects': N_subjects[roi].get(field, 0),
                })
            for field in sorted(our_data.get(roi, {})):
                writer.writerow({
                    'ROI': roi, 'source': 'present_study',
                    'field_T': field, 'mean_ppb': round(our_data[roi][field], 2),
                    'k_papers': 1, 'N_subjects': 7,
                })
    print(f"CSV  -> {csv_path}")


if __name__ == '__main__':
    main()
